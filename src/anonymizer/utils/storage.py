"""
This module provides utility functions for working with storage in the anonymizer application.

Functions:
- count_studies_series_images(patient_path: str) -> Tuple[int, int, int]: Counts the number of studies, series, and images in a given patient directory.
- count_study_images(base_dir: Path, anon_pt_id: str, study_uid: str) -> int: Counts the number of images stored in a given study directory.
- read_java_anonymizer_index_xlsx(filename: str) -> List[JavaAnonymizerExportedStudy]: Read data from the Java Anonymizer exported patient index file.

Classes:
- JavaAnonymizerExportedStudy: Represents the data structure for a single exported study from the Java Anonymizer.

"""

import os
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Union

import csv
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.child import _WorkbookChild
from openpyxl.worksheet.worksheet import Worksheet

from anonymizer.utils.translate import get_current_language_code

DICOM_FILE_SUFFIX = ".dcm"

logger = logging.getLogger(__name__)

def count_studies_series_images(patient_path: str) -> tuple[int, int, int]:
    """
    Counts the number of studies, series, and images in a given patient directory in the anonymizer store

    Args:
        patient_path (str): The path to the patient directory.

    Returns:
        A tuple containing the number of studies, series, and images in the patient directory.
    """
    study_count = 0
    series_count = 0
    image_count = 0

    for root, dirs, files in os.walk(patient_path):
        if root == patient_path:
            study_count += len(dirs)
        else:
            series_count += len(dirs)
        for file in files:
            if file.endswith(DICOM_FILE_SUFFIX):
                image_count += 1

    return study_count, series_count, image_count


def count_series(base_dir: str, patient_ids: Optional[list[str]] = None) -> int:
    """
    Counts the total number of series across multiple patients in the anonymizer store.

    Args:
        base_dir (str): The base directory containing patient folders.
        patient_ids (list[str]): A list of patient IDs, if None count number of series for ALL patients

    Returns:
        The total number of series across all patient directories in patient_id list
    """
    total_series = 0

    base_path = Path(base_dir)
    if not base_path.is_dir():
        raise ValueError(f"{base_dir} is not a valid directory")

    # If patient_ids not specified, iterate through ALL patients
    if patient_ids is None:
        patient_ids = [str(object=p) for p in base_path.iterdir() if p.is_dir()]

    for patient_id in patient_ids:
        patient_path: Path = base_path / patient_id
        if not patient_path.exists():
            continue

        for root, dirs, _ in os.walk(patient_path):
            if root != str(object=patient_path):  # Count series in subdirectories only
                total_series += len(dirs)

    return total_series


def get_dcm_files(root_path: str | Path) -> list[Path]:
    """
    Retrieves paths of each dicom file from a root path which could be at patient, study or series level.

    Args:
        root_path (str | Path): The root path to start the search for dicom files.

    Returns:
        List of Path objects
    """
    root_path = Path(root_path)  # Ensure root_path is a Path object

    return [
        Path(root) / file for root, _, files in os.walk(root_path) for file in files if file.endswith(DICOM_FILE_SUFFIX)
    ]


def count_study_images(base_dir: Path, anon_pt_id: str, study_uid: str) -> int:
    """
    Counts the number of images stored in a given study directory.

    Args:
        anon_uid (str): The anonymous patient ID.
        study_uid (str): The study UID.

    Returns:
        The number of images stored in the study directory.
    """
    study_path = Path(base_dir, anon_pt_id, study_uid)
    image_count = 0

    for _, _, files in os.walk(study_path):
        for file in files:
            if file.endswith(DICOM_FILE_SUFFIX):
                image_count += 1

    return image_count


@dataclass
class JavaAnonymizerExportedStudy:
    ANON_PatientName: str
    ANON_PatientID: str
    PHI_PatientName: str
    PHI_PatientID: str
    DateOffset: str
    ANON_StudyDate: str
    PHI_StudyDate: str
    ANON_Accession: str
    PHI_Accession: str
    ANON_StudyInstanceUID: str
    PHI_StudyInstanceUID: str


def read_java_anonymizer_index_xlsx(filename: str) -> list[JavaAnonymizerExportedStudy]:
    """
    Read data from the Java Anonymizer exported patient index file
    containing a single workbook & sheet with fields as per the JavaAnonymizerExportedStudy dataclass.

    Args:
        filename (str): The path to the Excel file.

    Returns:
        List of JavaAnonymizerExportedStudy dataclass objects.

    Raises:
        ValueError: If no active sheet is found in the workbook.
        FileNotFoundError: If the file is not found.

    If the sheet is empty, an empty list is returned.
    """

    workbook: Workbook = load_workbook(filename)
    sheet: Union[_WorkbookChild, None] = workbook.active
    data: list[JavaAnonymizerExportedStudy] = []

    if sheet is None or not isinstance(sheet, Worksheet):
        raise ValueError("No active sheet found in the workbook")

    for row in sheet.iter_rows(values_only=True, min_row=2):
        str_row = [str(item) if item is not None else "" for item in row]
        data.append(JavaAnonymizerExportedStudy(*str_row))

    return data


def default_whitelist_path(modality_code: str) -> Path:
    return Path(
        "assets/locales/" + str(get_current_language_code() or "en_US") + "/whitelists/" + modality_code + ".txt"
    )


def project_whitelist_path(project_dir: Path, modality_code: str) -> Path:
    return project_dir / Path("whitelists/" + modality_code + ".txt")


def load_whitelist_from_txt(filepath: Path) -> list[str]:
    whitelist = []
    if not filepath.is_file():
        raise ValueError(f"{filepath} is not a valid file")

    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            # 1. Remove comment part first (split at #, take first part)
            word_part = line.split("#", 1)[0]
            # 2. Strip whitespace from the word part
            word = word_part.strip()
            # 3. Add if not empty
            if word and word not in whitelist:
                whitelist.append(word.upper())

    return whitelist


def save_whitelist_to_txt(filepath: Path, whitelist: list[str]) -> None:
    with open(filepath, "w", encoding="utf-8") as f:
        for word in whitelist:
            f.write(word + "\n")


def save_project_whitelist(project_dir: Path, modality_code: str, whitelist: list[str]) -> Path:
    """
    Save the project whitelist to a text file in /project_dir/whitelists/modality_code.txt.

    Args:
        project_dir (Path): The main project directory
        modality_code (str): The modality code for which to save the whitelist.
        whitelist (list[str]): The whitelist to save.
    """
    if not whitelist:
        raise ValueError("Whitelist is empty")
    if not project_dir.is_dir():
        raise ValueError(f"{project_dir} is not a valid directory")

    filepath = project_whitelist_path(project_dir, modality_code)
    # Ensure the parent directory exists
    filepath.parent.mkdir(parents=True, exist_ok=True)
    # Save the whitelist to the file
    save_whitelist_to_txt(filepath, whitelist)
    return filepath


def load_project_whitelist(project_dir: Path, modality_code: str) -> list[str]:
    """
    Load the project whitelist for a given modality code.

    Args:
        project_dir (Path): The main project directory
        modality_code (str): The modality code for which to load the whitelist.

    Returns:
        A set of whitelisted terms.
    """
    return load_whitelist_from_txt(project_whitelist_path(project_dir, modality_code))


def load_default_whitelist(modality_code: str) -> list[str]:
    """
    Load the default whitelist for a given modality code.

    Args:
        modality_code (str): The modality code for which to load the whitelist.

    Returns:
        A set of whitelisted terms.
    """
    if modality_code == "CR" or modality_code == "MG":
        modality_code = "DX"
    return load_whitelist_from_txt(default_whitelist_path(modality_code))



def load_pseudo_keys(pseudo_key_path: Optional[Path]) -> tuple[dict[str, str], list[str]]:
    """
    Loads pseudo-anonymization keys from a .csv or .xlsx file.

    Args:
        pseudo_key_path (Optional[Path]): Path to the anonymization key file.

    Returns:
        tuple[dict[str, str], list[str]]: Mapping from original to anonymized IDs and a list of log messages.
    """
    messages = []
    mapping: dict[str, str] = {}

    if not pseudo_key_path:
        msg = "No anonymization key file specified, defaulting to automatic generation of anonymized patient IDs."
        logger.info(msg)
        messages.append(msg)
        return {}, messages

    if not pseudo_key_path.exists():
        msg = f"Pseudo Anonymization Key File not found: {pseudo_key_path}"
        logger.warning(msg)
        messages.append(msg)
        return {}, messages

    try:
        if pseudo_key_path.suffix.lower() == ".csv":
            mapping = _read_pseudo_mapping_csv(pseudo_key_path)
        elif pseudo_key_path.suffix.lower() == ".xlsx":
            mapping = _read_pseudo_mapping_xlsx(pseudo_key_path)
        else:
            msg = f"Unsupported file format: {pseudo_key_path.suffix}. Use '.csv' or '.xlsx'."
            logger.warning(msg)
            messages.append(msg)
    except ValueError as e:
        msg = str(e)
        logger.warning(msg)
        messages.append(msg)
    except Exception as e:
        msg = f"Unexpected error while loading pseudo key file: {pseudo_key_path}\n{e}"
        logger.exception(msg)
        messages.append(msg)

    return mapping, messages


def _read_pseudo_mapping_csv(path: Path) -> dict[str, str]:
    mapping = {}
    seen_orig = set()
    seen_anon = set()

    with path.open(newline="") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"CSV file '{path}' is empty.")

        indices = _detect_header_indices(header)
        if indices is None:
            raise ValueError("CSV file must contain recognizable column names for original and anonymized patient IDs. For example 'original patient id' and 'anonymous patient id'")

        orig_idx, anon_idx = indices

        for row in reader:
            if len(row) <= max(orig_idx, anon_idx):
                continue
            orig = row[orig_idx].strip()
            anon = row[anon_idx].strip()

            if orig in seen_orig:
                raise ValueError(f"Duplicate original patient ID found: '{orig}'. Key file can not contain duplicates.")
            if anon in seen_anon:
                raise ValueError(f"Duplicate anonymized patient ID found: '{anon}'. Key file can not contain duplicates.")

            if orig and anon:
                mapping[orig] = anon
                seen_orig.add(orig)
                seen_anon.add(anon)
    return mapping


def _read_pseudo_mapping_xlsx(path: Path) -> dict[str, str]:
    mapping = {}
    seen_orig = set()
    seen_anon = set()

    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError(f"XLSX file '{path}' is empty.")

    header = [str(cell).strip() if cell else "" for cell in rows[0]]
    indices = _detect_header_indices(header)
    if indices is None:
            raise ValueError("XLSX file must contain recognizable column names for original and anonymized patient IDs. For example 'original patient id' and 'anonymous patient id'")

    orig_idx, anon_idx = indices

    for row in rows[1:]:
        if len(row) <= max(orig_idx, anon_idx):
            continue
        orig = str(row[orig_idx]).strip() if row[orig_idx] else ""
        anon = str(row[anon_idx]).strip() if row[anon_idx] else ""

        if orig in seen_orig:
            raise ValueError(f"Duplicate original patient ID found: '{orig}'. Key file can not contain duplicates.")
        if anon in seen_anon:
            raise ValueError(f"Duplicate anonymized patient ID found: '{anon}'. Key file can not contain duplicates.")

        if orig and anon:
            mapping[orig] = anon
            seen_orig.add(orig)
            seen_anon.add(anon)
    return mapping


def _detect_header_indices(header: list[str]) -> tuple[int, int] | None:
    lowered = [h.lower().strip() for h in header]

    known_originals = {"original", "original id", "original patient id", "id"}
    known_anons = {"anon", "anonymous", "anonymized", "anonymous id", "anonymized id", "anonymous patient id", "anonymized patient id"}

    original_index = next((i for i, col in enumerate(lowered) if col in known_originals), -1)
    anon_index = next((i for i, col in enumerate(lowered) if col in known_anons), -1)

    if original_index == -1 or anon_index == -1:
        return None

    return original_index, anon_index

