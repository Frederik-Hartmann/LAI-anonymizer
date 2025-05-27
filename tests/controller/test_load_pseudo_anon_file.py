import csv
import pytest
from openpyxl import Workbook

from anonymizer.utils.storage import (
    load_pseudo_keys,
    _read_pseudo_mapping_csv,
    _read_pseudo_mapping_xlsx,
    _detect_header_indices,
)


# ------------------------------
# CSV Tests
# ------------------------------

def test_read_csv_valid(tmp_path):
    file = tmp_path / "test.csv"
    with file.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Original Patient ID", "Anonymized Patient ID"])
        writer.writerow(["001", "anon1"])
        writer.writerow(["002", "anon2"])

    result = _read_pseudo_mapping_csv(file)
    assert result == {"001": "anon1", "002": "anon2"}

def test_read_xlsx_valid_inverted(tmp_path):
    file = tmp_path / "test.csv"
    with file.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Anonymized Patient ID", "Original Patient ID"])
        writer.writerow(["anonA", "123"])
        writer.writerow(["anonB", "456"])

    result = _read_pseudo_mapping_xlsx(file)
    assert result == {"123": "anonA", "456": "anonB"}


def test_read_csv_invalid_header(tmp_path):
    file = tmp_path / "invalid.csv"
    with file.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Name", "Code"])
        writer.writerow(["001", "anon1"])

    with pytest.raises(ValueError, match="must contain recognizable column names"):
        _read_pseudo_mapping_csv(file)


def test_read_csv_empty(tmp_path):
    file = tmp_path / "empty.csv"
    file.write_text("")

    with pytest.raises(ValueError, match="is empty"):
        _read_pseudo_mapping_csv(file)


def test_read_csv_missing_data_row(tmp_path):
    file = tmp_path / "partial.csv"
    with file.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Original Patient ID", "Anonymized Patient ID"])
        writer.writerow(["only_one_value"])

    result = _read_pseudo_mapping_csv(file)
    assert result == {}

def test_csv_duplicate_original_id_raises(tmp_path):
    file = tmp_path / "partial.csv"
    with file.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Original Patient ID", "Anonymized Patient ID"])
        writer.writerow(["123", "abc"])
        writer.writerow(["123", "def"])

    with pytest.raises(ValueError, match="Duplicate original patient ID found: '123'"):
        _read_pseudo_mapping_csv(file)


def test_csv_duplicate_anonymized_id_raises(tmp_path):
    file = tmp_path / "partial.csv"
    with file.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Original Patient ID", "Anonymized Patient ID"])
        writer.writerow(["123", "abc"])
        writer.writerow(["345", "abc"])

    with pytest.raises(ValueError, match="Duplicate anonymized patient ID found: 'abc'"):
        _read_pseudo_mapping_csv(file)

# ------------------------------
# XLSX Tests
# ------------------------------

def test_read_xlsx_valid(tmp_path):
    file = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Original Patient ID", "Anonymized Patient ID"])
    ws.append(["123", "anonA"])
    ws.append(["456", "anonB"])
    wb.save(file)

    result = _read_pseudo_mapping_xlsx(file)
    assert result == {"123": "anonA", "456": "anonB"}

def test_read_xlsx_valid_inverted(tmp_path):
    file = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Anonymized Patient ID", "Original Patient ID"])
    ws.append(["anonA", "123"])
    ws.append(["anonB", "456"])
    wb.save(file)

    result = _read_pseudo_mapping_xlsx(file)
    assert result == {"123": "anonA", "456": "anonB"}


def test_read_xlsx_invalid_header(tmp_path):
    file = tmp_path / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Foo", "Bar"])
    ws.append(["123", "anon"])
    wb.save(file)

    with pytest.raises(ValueError, match="must contain recognizable column names"):
        _read_pseudo_mapping_xlsx(file)


def test_read_xlsx_empty_file(tmp_path):
    file = tmp_path / "empty.xlsx"
    wb = Workbook()
    wb.save(file)

    with pytest.raises(ValueError, match="is empty"):
        _read_pseudo_mapping_xlsx(file)


def test_read_xlsx_row_missing_data(tmp_path):
    file = tmp_path / "rowmiss.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Original Patient ID", "Anonymized Patient ID"])
    ws.append(["only_original"])
    wb.save(file)

    result = _read_pseudo_mapping_xlsx(file)
    assert result == {}


def test_xlsx_duplicate_original_id_raises(tmp_path):
    file = tmp_path / "rowmiss.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Original ID", "Anonymized ID"])
    ws.append(["123", "abc"])
    ws.append(["123", "xyz"])
    wb.save(file)

    with pytest.raises(ValueError, match="Duplicate original patient ID found: '123'"):
        _read_pseudo_mapping_xlsx(file)


def test_xlsx_duplicate_anonymized_id_raises(tmp_path):
    file = tmp_path / "rowmiss.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Original ID", "Anonymized ID"])
    ws.append(["123", "abc"])
    ws.append(["456", "abc"])
    wb.save(file)
    with pytest.raises(ValueError, match="Duplicate anonymized patient ID found: 'abc'"):
        _read_pseudo_mapping_xlsx(file)


# ------------------------------
# Header Detection Tests
# ------------------------------

@pytest.mark.parametrize("header,expected", [
    (["Original Patient ID", "Anonymized Patient ID"], (0, 1)),
    ([" original id ", " anonymized id "], (0, 1)),
    (["ORIGINAL", "ANONYMIZED"], (0, 1)),
    (["anon", "original"], (1, 0)),
    (["Name", "Code"], None),
])
def test_detect_header_indices(header, expected):
    assert _detect_header_indices(header) == expected


# ------------------------------
# Integration Tests
# ------------------------------

def test_load_pseudo_keys_csv(tmp_path):
    path = tmp_path / "keys.csv"
    with path.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Original Patient ID", "Anonymized Patient ID"])
        writer.writerow(["x", "y"])

    result, messages = load_pseudo_keys(path)
    assert result == {"x": "y"}


def test_load_pseudo_keys_xlsx(tmp_path):
    path = tmp_path / "keys.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Original Patient ID", "Anonymized Patient ID"])
    ws.append(["a", "b"])
    wb.save(path)

    result, messages = load_pseudo_keys(path)
    assert result == {"a": "b"}


def test_load_pseudo_keys_unsupported_format(tmp_path, caplog):
    file = tmp_path / "invalid.txt"
    file.write_text("Nonsense")
    result, messages = load_pseudo_keys(file)
    assert result == {}
    assert "Unsupported file format" in caplog.text


def test_load_pseudo_keys_missing_file(tmp_path, caplog):
    missing = tmp_path / "doesnotexist.csv"
    result, messages = load_pseudo_keys(missing)
    assert result == {}
    assert "not found" in caplog.text


def test_load_pseudo_keys_none_path(caplog):
    caplog.set_level("INFO")
    result, messages = load_pseudo_keys(None)
    assert result == {}
    assert "No anonymization key file specified" in caplog.text
